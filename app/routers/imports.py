"""Import de projets depuis Excel ou Word, et alimentation par les données collectées."""
from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, Form as FormField, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..crud import ensure_project, log_action, serialize
from ..database import get_db
from ..models import Form, FormQuestion, FormSubmission, Indicator, IndicatorActual, Project, User
from ..security import can_edit, can_manage
from ..services import importer
from ..services.xlsform import _nom_technique

router = APIRouter(prefix="/api/imports", tags=["Imports"])

TAILLE_MAX = 20 * 1024 * 1024  # 20 Mo


async def _lire(fichier: UploadFile) -> bytes:
    contenu = await fichier.read()
    if len(contenu) > TAILLE_MAX:
        raise HTTPException(status_code=413, detail="Fichier trop volumineux (20 Mo maximum).")
    if not contenu:
        raise HTTPException(status_code=422, detail="Le fichier reçu est vide.")
    return contenu


@router.post("/excel/{project_id}")
async def importer_excel(project_id: int, fichier: UploadFile = File(...),
                         remplacer: bool = FormField(False),
                         db: Session = Depends(get_db), user: User = Depends(can_edit)):
    """Charge un cadre logique complet (résultats, indicateurs, activités, budget, risques)
    depuis un classeur Excel conforme au modèle SEPIA ou à une structure proche."""
    projet = ensure_project(db, project_id)
    if not (fichier.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=422, detail="Format attendu : .xlsx")
    contenu = await _lire(fichier)
    try:
        rapport = importer.importer_excel(db, contenu, projet, remplacer=remplacer)
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=422,
                            detail=f"Lecture impossible : {type(exc).__name__} — {exc}")
    log_action(db, user, "IMPORT_EXCEL", "Project", projet.id, projet.id, str(rapport["crees"]))
    db.commit()
    return rapport


@router.post("/word/analyser")
async def analyser_word(fichier: UploadFile = File(...), user: User = Depends(can_edit)):
    """Inspecte un document Word et propose les tableaux exploitables."""
    if not (fichier.filename or "").lower().endswith(".docx"):
        raise HTTPException(status_code=422, detail="Format attendu : .docx")
    contenu = await _lire(fichier)
    try:
        return importer.analyser_word(contenu)
    except Exception as exc:
        raise HTTPException(status_code=422,
                            detail=f"Lecture impossible : {type(exc).__name__} — {exc}")


@router.post("/word/{project_id}")
async def importer_word(project_id: int, fichier: UploadFile = File(...),
                        index_tableau: Optional[int] = FormField(None),
                        db: Session = Depends(get_db), user: User = Depends(can_edit)):
    """Importe un cadre logique rédigé dans un document Word (matrice sous forme de tableau)."""
    projet = ensure_project(db, project_id)
    if not (fichier.filename or "").lower().endswith(".docx"):
        raise HTTPException(status_code=422, detail="Format attendu : .docx")
    contenu = await _lire(fichier)
    try:
        rapport = importer.importer_word(db, contenu, projet, index_tableau)
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=422,
                            detail=f"Import impossible : {type(exc).__name__} — {exc}")
    log_action(db, user, "IMPORT_WORD", "Project", projet.id, projet.id, str(rapport))
    db.commit()
    return rapport


@router.post("/kobo/{form_id}")
async def importer_reponses(form_id: int, fichier: UploadFile = File(...),
                            db: Session = Depends(get_db), user: User = Depends(can_edit)):
    """Réinjecte un export KoboToolbox/ODK (XLSX) dans la plateforme.

    Les colonnes portant le nom technique d'une question reliée à un indicateur
    alimentent automatiquement les réalisations de cet indicateur (agrégation par
    somme pour les effectifs, par moyenne pour les taux et scores).
    """
    form = db.get(Form, form_id)
    if not form:
        raise HTTPException(status_code=404, detail="Formulaire introuvable.")
    if not (fichier.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=422, detail="Format attendu : .xlsx")
    contenu = await _lire(fichier)
    try:
        classeur = load_workbook(BytesIO(contenu), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Classeur illisible : {exc}")
    feuille = classeur[classeur.sheetnames[0]]
    lignes = list(feuille.iter_rows(values_only=True))
    if len(lignes) < 2:
        raise HTTPException(status_code=422, detail="Le fichier ne contient aucune réponse.")
    entetes = [str(c) if c is not None else "" for c in lignes[0]]
    index_par_nom = {importer.normaliser(e): i for i, e in enumerate(entetes)}

    questions = {(_nom_technique(q.name or q.label)): q for q in form.questions}
    reponses = 0
    valeurs_par_indicateur: Dict[str, List[float]] = {}
    for ligne in lignes[1:]:
        if not any(c not in (None, "") for c in ligne):
            continue
        donnees: Dict[str, Any] = {}
        for nom, question in questions.items():
            index = index_par_nom.get(importer.normaliser(nom))
            if index is None or index >= len(ligne):
                continue
            valeur = ligne[index]
            donnees[nom] = valeur.isoformat() if hasattr(valeur, "isoformat") else valeur
            if question.linked_indicator_code:
                numerique = importer._valeur_numerique(valeur)
                if numerique is not None:
                    valeurs_par_indicateur.setdefault(
                        question.linked_indicator_code, []).append(numerique)
        db.add(FormSubmission(form_id=form_id, submitted_by=user.full_name, answers=donnees))
        reponses += 1

    indicateurs_alimentes = []
    periode = f"IMPORT-{fichier.filename[:20]}"
    for code, valeurs in valeurs_par_indicateur.items():
        indicateur = db.query(Indicator).filter(Indicator.project_id == form.project_id,
                                                Indicator.code == code).first()
        if not indicateur or not valeurs:
            continue
        agregation = "moyenne" if (indicateur.unit or "").strip() in ("%", "Score", "Ratio",
                                                                      "Indice") else "somme"
        valeur = round(sum(valeurs) / len(valeurs), 2) if agregation == "moyenne" else sum(valeurs)
        db.add(IndicatorActual(indicator_id=indicateur.id, period_label=periode,
                               value=valeur, source=f"Import {fichier.filename}",
                               collected_by=user.full_name, validation_status="Brouillon",
                               comment=f"Agrégation « {agregation} » sur {len(valeurs)} réponses."))
        indicateurs_alimentes.append({"code": code, "valeur": valeur, "agregation": agregation,
                                      "nb_reponses": len(valeurs)})
    log_action(db, user, "IMPORT_KOBO", "Form", form_id, form.project_id, f"{reponses} réponses")
    db.commit()
    return {"reponses_importees": reponses, "indicateurs_alimentes": indicateurs_alimentes,
            "periode_creee": periode if indicateurs_alimentes else None}


@router.post("/xlsform/{project_id}")
async def importer_xlsform(project_id: int, fichier: UploadFile = File(...),
                           db: Session = Depends(get_db), user: User = Depends(can_edit)):
    """Importe un XLSForm existant (feuilles survey/choices) comme formulaire SEPIA."""
    projet = ensure_project(db, project_id)
    contenu = await _lire(fichier)
    try:
        classeur = load_workbook(BytesIO(contenu), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Classeur illisible : {exc}")
    if "survey" not in [n.lower() for n in classeur.sheetnames]:
        raise HTTPException(status_code=422, detail="Le classeur ne contient pas de feuille « survey ».")
    nom_survey = next(n for n in classeur.sheetnames if n.lower() == "survey")
    lignes = list(classeur[nom_survey].iter_rows(values_only=True))
    entetes = [importer.normaliser(c) for c in lignes[0]]

    def colonne(nom: str) -> Optional[int]:
        return entetes.index(nom) if nom in entetes else None

    i_type, i_name, i_label = colonne("type"), colonne("name"), colonne("label")
    if i_type is None or i_name is None:
        raise HTTPException(status_code=422, detail="Feuille « survey » incomplète (type/name requis).")
    i_hint, i_required = colonne("hint"), colonne("required")
    i_relevant, i_constraint = colonne("relevant"), colonne("constraint")

    choix_par_liste: Dict[str, List[Dict[str, str]]] = {}
    nom_choices = next((n for n in classeur.sheetnames if n.lower() == "choices"), None)
    if nom_choices:
        lignes_choix = list(classeur[nom_choices].iter_rows(values_only=True))
        entetes_choix = [importer.normaliser(c) for c in lignes_choix[0]]
        c_liste = entetes_choix.index("list name") if "list name" in entetes_choix else 0
        c_nom = entetes_choix.index("name") if "name" in entetes_choix else 1
        c_libelle = entetes_choix.index("label") if "label" in entetes_choix else 2
        for ligne in lignes_choix[1:]:
            if not ligne or ligne[c_liste] is None:
                continue
            choix_par_liste.setdefault(str(ligne[c_liste]), []).append(
                {"name": str(ligne[c_nom]), "label": str(ligne[c_libelle])})

    form = Form(project_id=projet.id, name=fichier.filename.rsplit(".", 1)[0][:200],
                form_type="Questionnaire", code=f"IMP{projet.id}")
    db.add(form)
    db.flush()
    section = None
    position = 0
    for ligne in lignes[1:]:
        if not ligne or ligne[i_type] is None:
            continue
        type_brut = str(ligne[i_type]).strip()
        nom = str(ligne[i_name] or "").strip()
        libelle = str(ligne[i_label] or nom) if i_label is not None else nom
        if type_brut.startswith("begin"):
            section = libelle
            continue
        if type_brut.startswith("end") or type_brut in ("start", "end", "today", "deviceid",
                                                        "audit", "username", "simserial"):
            continue
        choix: List[Dict[str, str]] = []
        type_question = type_brut
        if type_brut.startswith(("select_one", "select_multiple")):
            morceaux = type_brut.split()
            type_question = morceaux[0]
            if len(morceaux) > 1:
                choix = choix_par_liste.get(morceaux[1], [])
        db.add(FormQuestion(
            form_id=form.id, order_index=position, section=section, name=nom[:80],
            label=libelle, question_type=type_question, choices=choix,
            required=str(ligne[i_required]).lower() in ("yes", "true", "1")
            if i_required is not None and ligne[i_required] else False,
            hint=str(ligne[i_hint]) if i_hint is not None and ligne[i_hint] else None,
            relevant=str(ligne[i_relevant]) if i_relevant is not None and ligne[i_relevant] else None,
            constraint=str(ligne[i_constraint]) if i_constraint is not None and ligne[i_constraint] else None,
        ))
        position += 1
    db.commit()
    db.refresh(form)
    return {"formulaire": serialize(form), "questions_importees": position}
